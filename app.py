import json
import os
import base64
import numpy as np
import pandas as pd
import requests
from sklearn.neighbors import KNeighborsRegressor
import streamlit as st
import copy

# Initialize session state variables
if "p_file_key" not in st.session_state:
    st.session_state.p_file_key = 0

if "cpt_file_key" not in st.session_state:
    st.session_state.cpt_file_key = 0

# Must be the absolute first Streamlit command in the script
st.set_page_config(page_title="Refrigerator Simulator Hub", layout="wide")

# =================================================================
# 1. GITHUB-BACKED PERSISTENCE LAYER
# =================================================================
# Local disk on Streamlit Cloud is wiped on every reboot/redeploy, so the
# database matrix is instead stored as a JSON file committed directly to
# your GitHub repo via the Contents API. Requires a [github] block in
# Streamlit secrets: token, repo ("user/repo"), and optionally branch.
REVIEWER_PASSWORD = "Admin@Cooling2026"

GITHUB_FILE_PATH = "simulator_storage.json"

def _github_config_ok():
    return "github" in st.secrets and "token" in st.secrets["github"] and "repo" in st.secrets["github"]

def _github_api_url():
    repo = st.secrets["github"]["repo"]
    return f"https://api.github.com/repos/{repo}/contents/{GITHUB_FILE_PATH}"

def _github_headers():
    token = st.secrets["github"]["token"]
    return {"Authorization": f"token {token}", "Accept": "application/vnd.github+json"}

def load_memory_from_disk():
    """Reads the saved database matrix from the GitHub repo on startup."""
    if not _github_config_ok():
        st.error("⚠️ GitHub storage isn't configured. Add a [github] block (token, repo) to Streamlit secrets.")
        return {}
    try:
        branch = st.secrets["github"].get("branch", "main")
        resp = requests.get(_github_api_url(), headers=_github_headers(), params={"ref": branch}, timeout=15)
        if resp.status_code == 200:
            content = resp.json()
            decoded = base64.b64decode(content["content"]).decode("utf-8")
            return json.loads(decoded) if decoded.strip() else {}
        elif resp.status_code == 404:
            # No file yet — first run. It will be created on first save.
            return {}
        else:
            st.error(f"⚠️ GitHub load error ({resp.status_code}): {resp.text}")
            return {}
    except Exception as e:
        st.error(f"⚠️ Error loading backup database file from GitHub: {str(e)}")
        return {}

def save_memory_to_disk(db_matrix):
    """Commits the current database matrix to the GitHub repo instantly."""
    if not _github_config_ok():
        st.error("⚠️ GitHub storage isn't configured. Add a [github] block (token, repo) to Streamlit secrets.")
        return
    try:
        branch = st.secrets["github"].get("branch", "main")
        headers = _github_headers()

        # Need the current file's SHA to update it (GitHub requires this for existing files)
        get_resp = requests.get(_github_api_url(), headers=headers, params={"ref": branch}, timeout=15)
        sha = get_resp.json().get("sha") if get_resp.status_code == 200 else None

        content_str = json.dumps(db_matrix, indent=4)
        encoded_content = base64.b64encode(content_str.encode("utf-8")).decode("utf-8")

        payload = {
            "message": "Update simulator storage data",
            "content": encoded_content,
            "branch": branch,
        }
        if sha:
            payload["sha"] = sha

        put_resp = requests.put(_github_api_url(), headers=headers, json=payload, timeout=15)
        if put_resp.status_code not in (200, 201):
            st.error(f"⚠️ Failed to write backup data to GitHub ({put_resp.status_code}): {put_resp.text}")
    except Exception as e:
        st.error(f"⚠️ Failed to write backup data to GitHub: {str(e)}")

# Legacy compatibility wrapper in case it's explicitly called later in your file
def save_memory(data):
    save_memory_to_disk(data)

# =================================================================
# 2. STATE INITIALIZATION ON STARTUP
# =================================================================
# --- Initialize Session States ---
if "db" not in st.session_state:
    st.session_state.db = load_memory_from_disk()

if 'reviewer_logged_in' not in st.session_state:
    st.session_state.reviewer_logged_in = False

if "arr_form_id" not in st.session_state:
    st.session_state.arr_form_id = 0

if "model_form_id" not in st.session_state:
    st.session_state.model_form_id = 0

# Global configuration constants
tc_features = ['tf-1', 'tf-2', 'tf-3', 'tf-4', 'tf-5', 'tc-1', 'tc-2', 'tc-3', 'tvc', 'S2']
metric_types = ['mean', 'min', 'max', '(max+min)/2']
metric_labels = {'mean': 'Mean', 'min': 'Min', 'max': 'Max', '(max+min)/2': '(Max+Min)/2'}

# =================================================================
# 3. UTILITY HELPER FUNCTIONS
# =================================================================
def normalize_sensor_name(name):
    if not isinstance(name, str):
        return ""
    # Lowercase, remove hyphens, spaces, and underscores (e.g., "tf-1" -> "tf1")
    return name.lower().replace(" ", "").replace("-", "").replace("_", "").strip()
def to_float(v):
    try:
        if pd.isna(v) or str(v).strip() == "":
            return 0.0
        return round(float(v), 1)
    except:
        return 0.0

# Rounds every numeric column in a dataframe to 1 decimal place for display/export/save
def round_df(df, decimals=1):
    df = df.copy()
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    if len(numeric_cols) > 0:
        df[numeric_cols] = df[numeric_cols].round(decimals)
    return df

# Builds a column_config dict: numeric columns get 1-decimal NumberColumns, text_cols
# get TextColumns, and any column named in disabled_cols is locked from editing.
def build_column_config(df, disabled_cols=None, text_cols=None):
    disabled_cols = set(disabled_cols or [])
    text_cols = set(text_cols or [])
    config = {}
    for col in df.columns:
        if col in text_cols:
            config[col] = st.column_config.TextColumn(col, disabled=(col in disabled_cols))
        else:
            config[col] = st.column_config.NumberColumn(col, format="%.1f", disabled=(col in disabled_cols))
    return config

# Blanks the Test Flag on continuation rows of a group (rows 2..N of each level),
# so the editable grid visually reads like the Test Flag cell is "merged" downward.
def blank_repeated_flag(df):
    df = df.copy()
    prev = object()  # sentinel that can't equal any real flag value
    for idx in df.index:
        cur = df.at[idx, "Test Flag"]
        if cur == prev:
            df.at[idx, "Test Flag"] = ""
        else:
            prev = cur
    return df

# Reverses blank_repeated_flag: fills blanked Test Flag cells with the nearest
# non-blank value above them, so downstream save/comparison logic sees full values.
def forward_fill_flag(df):
    df = df.copy()
    last = None
    filled = []
    for v in df["Test Flag"]:
        if v is None or (isinstance(v, str) and v.strip() == ""):
            filled.append(last)
        else:
            last = v
            filled.append(v)
    df["Test Flag"] = filled
    return df

# Renders a CPT dataframe as an HTML table with Test Flag, S2, and Sensor cells
# genuinely merged (rowspan) across each level's group of Mean/Min/Max/(Max+Min)/2 rows.
# Used for the read-only "Original Uploaded CPT Matrix" — st.data_editor can't merge cells,
# but a plain HTML table (for display only) can.
def render_merged_cpt_table(df):
    if df.empty:
        return "<p><em>No data</em></p>"

    df = df.reset_index(drop=True)
    cols = list(df.columns)
    merge_cols = {"Test Flag", "S2", "Sensor"}
    border = "border:1px solid rgba(120,120,120,0.5);"

    def fmt(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        if isinstance(v, (int, float, np.floating, np.integer)):
            return f"{float(v):.1f}"
        return str(v)

    html = ["<div style='overflow-x:auto;'>", "<table style='width:100%; border-collapse:collapse; font-size:14px;'>", "<thead><tr>"]
    for c in cols:
        html.append(f"<th style='{border} padding:6px 10px; text-align:center; font-weight:600;'>{c}</th>")
    html.append("</tr></thead><tbody>")

    n = len(df)
    i = 0
    while i < n:
        flag_val = df.loc[i, "Test Flag"]
        j = i
        while j < n and df.loc[j, "Test Flag"] == flag_val:
            j += 1
        group_size = j - i

        # Find the one row in this group that actually carries S2 (Mean row) / Sensor (Min row)
        s2_val, sensor_val = None, None
        for k in range(i, j):
            if "S2" in df.columns and pd.notna(df.loc[k, "S2"]):
                s2_val = df.loc[k, "S2"]
            if "Sensor" in df.columns and pd.notna(df.loc[k, "Sensor"]):
                sensor_val = df.loc[k, "Sensor"]

        for local_idx, row_idx in enumerate(range(i, j)):
            html.append("<tr>")
            for c in cols:
                if c == "Test Flag":
                    if local_idx == 0:
                        html.append(f"<td rowspan='{group_size}' style='{border} padding:6px 10px; text-align:center; vertical-align:middle; font-weight:600;'>{flag_val}</td>")
                elif c == "S2":
                    if local_idx == 0:
                        html.append(f"<td rowspan='{group_size}' style='{border} padding:6px 10px; text-align:center; vertical-align:middle;'>{fmt(s2_val)}</td>")
                elif c == "Sensor":
                    if local_idx == 0:
                        html.append(f"<td rowspan='{group_size}' style='{border} padding:6px 10px; text-align:center; vertical-align:middle;'>{fmt(sensor_val)}</td>")
                else:
                    html.append(f"<td style='{border} padding:6px 10px; text-align:center;'>{fmt(df.loc[row_idx, c])}</td>")
            html.append("</tr>")
        i = j

    html.append("</tbody></table></div>")
    return "".join(html)

# Computes tf-a (avg of tf-1..tf-5) and tc-a (avg of tc-1..tc-3) for a dict of sensor values
def compute_avg_fields(values_dict):
    tf_vals = [to_float(values_dict.get(f"tf-{i}", 0.0)) for i in range(1, 6)]
    tc_vals = [to_float(values_dict.get(f"tc-{i}", 0.0)) for i in range(1, 4)]
    tf_a = round(sum(tf_vals) / len(tf_vals), 1) if tf_vals else 0.0
    tc_a = round(sum(tc_vals) / len(tc_vals), 1) if tc_vals else 0.0
    return tf_a, tc_a

# Adds tf-a and tc-a average columns to a single-row-per-record dataframe,
# positioned immediately after tf-5 and tc-3 respectively
def add_avg_columns(df):
    df = df.copy()
    tf_cols = [c for c in ["tf-1", "tf-2", "tf-3", "tf-4", "tf-5"] if c in df.columns]
    tc_cols = [c for c in ["tc-1", "tc-2", "tc-3"] if c in df.columns]
    if tf_cols:
        df["tf-a"] = df[tf_cols].apply(pd.to_numeric, errors="coerce").mean(axis=1).round(1)
    if tc_cols:
        df["tc-a"] = df[tc_cols].apply(pd.to_numeric, errors="coerce").mean(axis=1).round(1)

    # Reorder so tf-a sits right after tf-5, and tc-a sits right after tc-3
    cols = list(df.columns)

    def move_after(cols, col_to_move, after_col):
        if col_to_move in cols and after_col in cols:
            cols.remove(col_to_move)
            cols.insert(cols.index(after_col) + 1, col_to_move)
        return cols

    cols = move_after(cols, "tf-a", "tf-5")
    cols = move_after(cols, "tc-a", "tc-3")
    return df[cols]

# Helper initialization to safeguard nested multi-arrangement structure
def verify_db_structure(vol, arr_name, p_amb, c_amb):
    if vol not in st.session_state.db or not isinstance(st.session_state.db[vol], dict):
        st.session_state.db[vol] = {}
    
    # Catch old layouts or non-existent arrangements
    if arr_name not in st.session_state.db[vol] or not isinstance(st.session_state.db[vol][arr_name], dict):
        st.session_state.db[vol][arr_name] = {}
        
    if p_amb not in st.session_state.db[vol][arr_name] or not isinstance(st.session_state.db[vol][arr_name][p_amb], dict):
        st.session_state.db[vol][arr_name][p_amb] = {}
        
    if c_amb not in st.session_state.db[vol][arr_name][p_amb]:
        st.session_state.db[vol][arr_name][p_amb][c_amb] = []

# =================================================================
# 4. ADVANCED INTERPOLATION ENGINE
# =================================================================
def run_automated_simulation(volume_records, new_pulldown, target_sensors):
    """
    Generates a consolidated prediction dataframe where each Sensor Query Point produces
    4 rows — Mean, Min, Max, (Max+Min)/2 — each predicted from that criteria's historical
    data. S2 is only meaningful on the Mean row (it was only ever recorded there).
    """
    if not volume_records:
        return pd.DataFrame()

    def clean_val(v):
        if v is None or pd.isna(v):
            return 0.0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    base_fields = ["tf-1", "tf-2", "tf-3", "tf-4", "tf-5", "tc-1", "tc-2", "tc-3", "tvc"]

    predicted_rows = []

    # Process each user manual query point, once per criteria (Mean/Min/Max/(Max+Min)/2)
    for target_s in target_sensors:
        for metric_key in metric_types:
            X_train = []
            y_train = []

            for record in volume_records:
                if 'cpt_data' not in record or not record['cpt_data']:
                    continue

                # Pulldown base features (unchanged — pulldown telemetry is still one flat vector)
                p_features = [clean_val(record["pulldown_data"].get(f, 0.0)) for f in tc_features]
                p_baseline = clean_val(record.get("pulldown_baseline_sensor", 0.0))

                # Treat each flag/level as a distinct data point to capture variance across sensor points
                for flag, level_data in record["cpt_data"].items():
                    metric_data = level_data.get(metric_key) if isinstance(level_data, dict) else None
                    if not metric_data:
                        continue

                    # Extract historical sensor cut-out temperature for this level block (flag-level, from the Min row)
                    hist_sensor = clean_val(level_data.get("Sensor", level_data.get("sensor", 0.0)))

                    # If Sensor value is missing or zero, use a default fallback matching common levels
                    if hist_sensor == 0.0:
                        if "1" in flag: hist_sensor = -27.5
                        elif "2" in flag: hist_sensor = -27.0
                        elif "3" in flag: hist_sensor = -25.5
                        elif "4" in flag: hist_sensor = -24.0
                        elif "5" in flag: hist_sensor = -21.0

                    # Formulate training input: Pulldown context matrix + specific level's target sensor value
                    train_input = p_features + [p_baseline, hist_sensor]

                    # Target output values to be predicted for this criteria (tf-1..5, tc-1..3, tvc)
                    train_target = [clean_val(metric_data.get(f, 0.0)) for f in base_fields]
                    # S2 was only ever recorded on the Mean row, so only train/predict it there
                    if metric_key == "mean":
                        train_target = train_target + [clean_val(level_data.get("S2", 0.0))]

                    X_train.append(train_input)
                    y_train.append(train_target)

            if len(X_train) >= 1:
                X_arr = np.nan_to_num(np.array(X_train, dtype=np.float64), nan=0.0)
                y_arr = np.nan_to_num(np.array(y_train, dtype=np.float64), nan=0.0)

                # Construct active query: current telemetry + the specific target query sensor step
                current_baseline = clean_val(new_pulldown[-1]) if len(new_pulldown) == 10 else 35.0
                query_vector = np.nan_to_num(
                    np.array([clean_val(v) for v in new_pulldown] + [current_baseline, clean_val(target_s)], dtype=np.float64).reshape(1, -1),
                    nan=0.0
                )

                # Run distance-weighted interpolation to compute dynamic shift across query points
                knn = KNeighborsRegressor(n_neighbors=min(3, len(X_arr)), weights='distance')
                knn.fit(X_arr, y_arr)

                prediction = knn.predict(query_vector)[0]

                row = {
                    "Sensor Value": f"{target_s} °C",
                    "Metric": metric_labels[metric_key],
                    "tf-1": round(prediction[0], 1),
                    "tf-2": round(prediction[1], 1),
                    "tf-3": round(prediction[2], 1),
                    "tf-4": round(prediction[3], 1),
                    "tf-5": round(prediction[4], 1),
                    "tc-1": round(prediction[5], 1),
                    "tc-2": round(prediction[6], 1),
                    "tc-3": round(prediction[7], 1),
                    "tvc": round(prediction[8], 1),
                    "S2": round(prediction[9], 1) if metric_key == "mean" else np.nan
                }
                predicted_rows.append(row)

    if predicted_rows:
        return pd.DataFrame(predicted_rows)
    return pd.DataFrame()


# =================================================================
# 5. STREAMLIT USER INTERFACE (TABS) & REPOSITORY LOGIC
# =================================================================

# ================= SIDEBAR: PROFILE & ARRANGEMENT MANAGER =================
st.sidebar.header("📁 Volume Profile Manager")

# Clean up empty strings or accidental "None" literal string keys if they exist in memory
if "None" in st.session_state.db:
    del st.session_state.db["None"]
if "" in st.session_state.db:
    del st.session_state.db[""]

existing_volumes = list(st.session_state.db.keys())
if existing_volumes:
    existing_volumes.sort(reverse=False)

# Render selectbox with actual available volumes
selected_volume = st.sidebar.selectbox(
    "Active Refrigerator Model:", 
    existing_volumes if existing_volumes else ["None"]
)

# --- Delete Model Option ---
if existing_volumes:
    st.sidebar.caption("⚠️ Permanently removes this model and all its arrangements")
    if st.sidebar.button("🗑️ Delete Selected Model"):
        # Allow deletion if there is at least one other valid model left
        if len(existing_volumes) > 1:
            if selected_volume in st.session_state.db:
                del st.session_state.db[selected_volume]
                save_memory(st.session_state.db)
                st.sidebar.success(f"Model '{selected_volume}' deleted successfully.")
                st.rerun()
        else:
            st.sidebar.error("❌ Cannot delete the last remaining model. Create another model before deleting this one.")

st.sidebar.markdown("---")

# 2. Select Arrangement of Selected Volume & Deletion
existing_arrangements = []
if selected_volume and selected_volume in st.session_state.db and isinstance(st.session_state.db[selected_volume], dict):
    existing_arrangements = list(st.session_state.db[selected_volume].keys())

if existing_arrangements:
    existing_arrangements.sort(reverse=False)
    selected_arrangement = st.sidebar.selectbox("Select Arrangement of Selected Volume:", existing_arrangements)
    
    # --- Delete Arrangement Option ---
    st.sidebar.caption("⚠️ Removes this arrangement data only")
    if st.sidebar.button("🗑️ Delete Selected Arrangement"):
        if len(existing_arrangements) > 1:
            del st.session_state.db[selected_volume][selected_arrangement]
            save_memory(st.session_state.db)
            st.sidebar.success(f"Arrangement '{selected_arrangement}' deleted.")
            st.rerun()
        else:
            st.sidebar.error("❌ Cannot delete the last remaining arrangement. Create another arrangement before deleting it.")
else:
    selected_arrangement = "None"
    st.sidebar.caption("No arrangements found. Register one below.")

st.sidebar.markdown("---")

# 3. ➕ Create New Arrangement Inputs
st.sidebar.subheader("📐 Design Arrangements")
new_arr = st.sidebar.text_input("➕ Create New Arrangement:", placeholder="e.g., A2", key=f"input_arr_{st.session_state.arr_form_id}")

if st.sidebar.button("Register Arrangement"):
    if new_arr and selected_volume and selected_volume != "None":
        new_arr_clean = new_arr.strip()
        if selected_volume not in st.session_state.db or not isinstance(st.session_state.db[selected_volume], dict):
            st.session_state.db[selected_volume] = {}
        if new_arr_clean not in st.session_state.db[selected_volume]:
            st.session_state.db[selected_volume][new_arr_clean] = {}
            save_memory(st.session_state.db)
            st.sidebar.success(f"Arrangement '{new_arr_clean}' registered under {selected_volume}!")
            st.session_state.arr_form_id += 1
            st.rerun()

st.sidebar.markdown("---")

# 4. ➕ Register New Volume Model Form
st.sidebar.subheader("➕ Register New Volume Model")
new_vol = st.sidebar.text_input("New Model Name:", placeholder="e.g., 365L", key=f"input_vol_{st.session_state.model_form_id}")
initial_arr = st.sidebar.text_input("Initial Arrangement Name:", placeholder="e.g., A1", key=f"input_init_{st.session_state.model_form_id}")

if st.sidebar.button("Add Volume Segment"):
    if new_vol and initial_arr:
        new_vol_clean = new_vol.strip()
        initial_arr_clean = initial_arr.strip()
        
        if new_vol_clean not in st.session_state.db:
            st.session_state.db[new_vol_clean] = {initial_arr_clean: {}}
            save_memory(st.session_state.db)
            st.success(f"Model {new_vol_clean} initialized with arrangement {initial_arr_clean}!")
            st.session_state.model_form_id += 1
            st.rerun()
        else:
            st.sidebar.error("This model name already exists.")
    elif new_vol or initial_arr:
        st.sidebar.error("⚠️ You must provide both the Model Name AND the Initial Arrangement Name.")


# === RESTORED TAB DEFINITIONS ===
tab1, tab2, tab3 = st.tabs(["🎛️ Run Automated Simulator", "🛠️ Data Repository Room", "🔍 Reviewer Dashboard"])

# ================= TAB 1: RUN AUTOMATED SIMULATOR =================
with tab1:
    st.subheader(f"Predict Multilevel CPT Matrices for [{selected_volume}] ({selected_arrangement})")
    
    c1, c2 = st.columns(2)
    with c1:
        sim_p_ambient = st.selectbox("Select Target Pulldown Ambient:", ["32°C", "43°C"], key="sim_p_amb")
    with c2:
        sim_c_ambient = st.selectbox("Select Target Respected CPT Ambient:", ["16°C", "32°C", "43°C"], key="sim_c_amb")
        
    p_key = "32C" if "32" in sim_p_ambient else "43C"
    c_key = "16C" if "16" in sim_c_ambient else ("32C" if "32" in sim_c_ambient else "43C")
    
    verify_db_structure(selected_volume, selected_arrangement, p_key, c_key)
    vol_records = st.session_state.db[selected_volume][selected_arrangement][p_key][c_key]
    
    # Initialize version counters to force-refresh fields upon upload
    if "sim_ver" not in st.session_state:
        st.session_state.sim_ver = 0
    if 'active_pulldown_form' not in st.session_state:
        st.session_state.active_pulldown_form = {}
        st.session_state.last_uploaded_sim_file = None

    if not vol_records:
        st.warning(f"⚠️ No matching profiles found under arrangement **{selected_arrangement}** for **Pulldown: {sim_p_ambient}** linked to **CPT: {sim_c_ambient}**.")
    else:
        # ================= STEP 1: AUTOMATED SIMULATOR PARSER =================
        st.markdown("#### Step 1: Input Current Pulldown Telemetry Vector")
        sim_pulldown_file = st.file_uploader(
            f"Auto-fill fields from local Pulldown Report ({sim_p_ambient})", 
            type=["xlsx", "xls"], key=f"sim_file_upload_{p_key}_{c_key}"
        )

        if sim_pulldown_file and sim_pulldown_file != st.session_state.last_uploaded_sim_file:
            try:
                # Read first available sheet dynamically
                df_sim_p = pd.read_excel(sim_pulldown_file, sheet_name=0, header=None)
                df_sim_p[0] = df_sim_p[0].astype(str).apply(normalize_sensor_name)
                
                sheet_data = {}
                for _, row in df_sim_p.dropna(subset=[0]).iterrows():
                    lbl = row[0]
                    if lbl not in sheet_data:
                        try:
                            sheet_data[lbl] = round(float(row[1]), 1)
                        except (ValueError, TypeError):
                            continue

                mapping_keys = {
                    'tf-1':'tf1', 'tf-2':'tf2', 'tf-3':'tf3', 'tf-4':'tf4', 'tf-5':'tf5', 
                    'tc-1':'tc1', 'tc-2':'tc2', 'tc-3':'tc3', 'S2':'s2'
                }
                
                # Extract individual thermocouple cells
                for feat, normalized_label in mapping_keys.items():
                    if normalized_label in sheet_data:
                        st.session_state.active_pulldown_form[feat] = sheet_data[normalized_label]
                        
                # Compute average for tvc from subcomponents if available
                tvc_values = [sheet_data[lbl] for lbl in ['tvc1', 'tvc2', 'tvc3'] if lbl in sheet_data]
                if tvc_values:
                    st.session_state.active_pulldown_form['tvc'] = round(sum(tvc_values) / len(tvc_values), 1)
                elif 'tvc' in sheet_data:
                    st.session_state.active_pulldown_form['tvc'] = sheet_data['tvc']

                st.session_state.last_uploaded_sim_file = sim_pulldown_file
                # Increment key version to instantly clear old component cache and force update UI inputs
                st.session_state.sim_ver += 1
                st.toast("🟢 Parsed values pulled from sheet!", icon="📊")
                st.rerun()
            except Exception as e:
                st.error(f"Error parsing configuration: {str(e)}")

        u_cols = st.columns(10)
        new_pulldown_input = []
        default_defaults = {'tf-1': -24.4, 'tf-2': -21.8, 'tf-3': -22.8, 'tf-4': -26.2, 'tf-5': -26.4, 'tc-1': 1.9, 'tc-2': 1.6, 'tc-3': 0.5, 'tvc': 8.1, 'S2': 41.1}
        
        for i, feat in enumerate(tc_features):
            # Prioritize extracted file data if available, otherwise use defaults
            if feat in st.session_state.active_pulldown_form:
                curr_val = round(float(st.session_state.active_pulldown_form[feat]), 1)
            else:
                curr_val = default_defaults.get(feat, 0.0)
                
            # Bound dynamic widget version to key parameters to force a redraw when new files parse
            val = u_cols[i].number_input(
                f"{feat}:", 
                value=curr_val, 
                step=0.1,
                format="%.1f",
                key=f"sim_inp_{p_key}_{c_key}_{feat}_v{st.session_state.sim_ver}"
            )
            new_pulldown_input.append(val)
            
        st.markdown("---")
        
        # ================= STEP 2: SET MULTI-SENSOR SIMULATION STEPS =================
        st.markdown("#### Step 2: Set Multi-Sensor Simulation Steps")
        
        num_targets = st.number_input("Number of target sensor points (1 to 7):", min_value=1, max_value=7, value=5, step=1)
        
        target_sensors = []
        s_cols = st.columns(int(num_targets))
        for idx in range(int(num_targets)):
            # Set up default sensor values for the 5 points (adjust the start value or step as needed)
            default_sensor_val = -27.5 + (idx * 1.5) 
            
            s_val = s_cols[idx].number_input(
                f"Sensor Query Point {idx+1} (°C):", 
                value=default_sensor_val, 
                step=0.1,
                format="%.1f",
                key=f"q_s_{p_key}_{c_key}_{idx}"
            )
            target_sensors.append(s_val)
            
        if st.button("🚀 Generate Predictive CPT Dataset Matrices", type="primary"):
            with st.spinner("Processing automated interpolation runs..."):
                df_final_predictions = run_automated_simulation(vol_records, new_pulldown_input, target_sensors)
                
                if df_final_predictions.empty:
                    st.error("Simulation engine run failed. Make sure dataset memory contains recorded instances.")
                else:
                    st.markdown("### 📊 Consolidated Predictive Simulation Output Matrix")
                    numeric_cols = [c for c in df_final_predictions.columns if c not in ("Sensor Value", "Metric")]
                    pred_col_config = {c: st.column_config.NumberColumn(c, format="%.1f") for c in numeric_cols}
                    st.dataframe(df_final_predictions, use_container_width=True, hide_index=True, column_config=pred_col_config)

# ================= TAB 2: DATA REPOSITORY ROOM =================
with tab2:
    st.subheader(f"Onboard Lab Reports for [{selected_volume}] ({selected_arrangement})")
    
    repo_c1, repo_c2 = st.columns(2)
    with repo_c1:
        repo_p_ambient = st.selectbox("Source Pulldown File Ambient Layer:", ["32°C", "43°C"], key="repo_p_amb")
    with repo_c2:
        repo_c_ambient = st.selectbox("Source Connected CPT File Ambient Layer:", ["16°C", "32°C", "43°C"], key="repo_c_amb")
        
    p_repo_key = "32C" if "32" in repo_p_ambient else "43C"
    c_repo_key = "16C" if "16" in repo_c_ambient else ("32C" if "32" in repo_c_ambient else "43C")
    
    col_f1, col_f2 = st.columns(2)
    with col_f1:
        repo_pulldown_file = st.file_uploader(
            f"Upload Pulldown Excel ({repo_p_ambient})", 
            type=["xlsx", "xls"], 
            key=f"r_p_file_{p_repo_key}_{c_repo_key}_v_{st.session_state.p_file_key}"
        )
    with col_f2:
        repo_cpt_file = st.file_uploader(
            f"Upload Respected CPT Excel ({repo_c_ambient})", 
            type=["xlsx", "xls"], 
            key=f"r_cpt_file_{p_repo_key}_{c_repo_key}_v_{st.session_state.cpt_file_key}"
        )
        
    if repo_pulldown_file and repo_cpt_file:
        if st.button("💾 Process and Train Simulator Memory Buffer", type="primary"):
            try:
                # 1. PARSE PULLDOWN DATA
                df_p_sum = pd.read_excel(repo_pulldown_file, sheet_name=0, header=None)
                df_p_sum[0] = df_p_sum[0].astype(str).apply(normalize_sensor_name)
                
                sheet_data = {}
                for _, row in df_p_sum.dropna(subset=[0]).iterrows():
                    lbl = row[0]
                    if lbl not in sheet_data:
                        try: sheet_data[lbl] = round(float(row[1]), 1)
                        except (ValueError, TypeError): continue

                mapping_keys = {
                    'tf-1':'tf1', 'tf-2':'tf2', 'tf-3':'tf3', 'tf-4':'tf4', 'tf-5':'tf5', 
                    'tc-1':'tc1', 'tc-2':'tc2', 'tc-3':'tc3', 'S2':'s2'
                }
                
                p_extracted = {}
                for target_key, normalized_label in mapping_keys.items():
                    p_extracted[target_key] = sheet_data.get(normalized_label, 0.0)
                        
                tvc_values = [sheet_data[lbl] for lbl in ['tvc1', 'tvc2', 'tvc3'] if lbl in sheet_data]
                if tvc_values:
                    p_extracted['tvc'] = round(sum(tvc_values) / len(tvc_values), 1)
                else:
                    p_extracted['tvc'] = 0.0
                    
                resolved_sensor = sheet_data.get('sensor', 0.0)
                
                # 2. PARSE CPT DATA
                cpt_structured = {}
                parsed_successfully = False

                # -------------------------------------------------------------
                # STRATEGY A : Visible Ambient Parser (FIXED VARIABLES)
                # -------------------------------------------------------------
                try:
                    import openpyxl
                    from openpyxl.utils import get_column_letter

                    wb = openpyxl.load_workbook(repo_cpt_file, data_only=True)
                    
                    if "ANALYSIS REPORT" in wb.sheetnames:
                        sheet_name = "ANALYSIS REPORT"
                    elif "CPT CALCULATION REPORT" in wb.sheetnames:
                        sheet_name = "CPT CALCULATION REPORT"
                    else:
                        sheet_name = wb.sheetnames[0]
                        
                    ws = wb[sheet_name]
                    cpt_structured = {}

                    # Safe numeric extraction, rounded to 1 decimal at the source
                    def safe_float(val):
                        if val is None:
                            return 0.0
                        try:
                            if isinstance(val, str):
                                val = val.replace("°C", "").replace("̊C", "").strip()
                            return round(float(val), 1)
                        except (ValueError, TypeError):
                            return 0.0

                    # Unhidden cell detection validation
                    def get_visible_value(row_idx, col_idx):
                        if ws.row_dimensions[row_idx].hidden:
                            return None
                        col_letter = get_column_letter(col_idx)
                        if ws.column_dimensions[col_letter].hidden:
                            return None
                        return ws.cell(row_idx, col_idx).value

                    start_row = None
                    for check_row in range(1, ws.max_row + 1):
                        if ws.row_dimensions[check_row].hidden:
                            continue
                        
                        raw_a = get_visible_value(check_row, 1)
                        raw_b = get_visible_value(check_row, 2)

                        a_str = str(raw_a).strip().lower() if raw_a is not None else ""
                        b_str = str(raw_b).strip().lower() if raw_b is not None else ""

                        if ("th. knob" in a_str or "th knob" in a_str) and "data criteria" in b_str:
                            start_row = check_row + 2
                            break

                    if start_row is None:
                        raise Exception("Visible CPT table header coordinates not found.")

                    current_flag = None
                    for data_row in range(start_row, ws.max_row + 1):
                        if ws.row_dimensions[data_row].hidden:
                            continue

                        colA = get_visible_value(data_row, 1)
                        colB = get_visible_value(data_row, 2)

                        colA = "" if colA is None else str(colA).strip()
                        colB = "" if colB is None else str(colB).strip().lower()

                        if "level" in colA.lower() or "boost" in colA.lower():
                            current_flag = colA
                            if current_flag not in cpt_structured:
                                cpt_structured[current_flag] = {"S2": 0.0, "Sensor": 0.0}

                        if current_flag is None:
                            continue
                        if current_flag not in cpt_structured:
                            cpt_structured[current_flag] = {"S2": 0.0, "Sensor": 0.0}

                        # Read strictly from mapped unhidden cells, one sub-block per criteria row
                        if colB in metric_types:
                            cpt_structured[current_flag][colB] = {
                                "tf-1": safe_float(get_visible_value(data_row, 3)),
                                "tf-2": safe_float(get_visible_value(data_row, 4)),
                                "tf-3": safe_float(get_visible_value(data_row, 5)),
                                "tf-4": safe_float(get_visible_value(data_row, 6)),
                                "tf-5": safe_float(get_visible_value(data_row, 7)),
                                "tc-1": safe_float(get_visible_value(data_row, 13)),
                                "tc-2": safe_float(get_visible_value(data_row, 14)),
                                "tc-3": safe_float(get_visible_value(data_row, 15)),
                                "tvc": safe_float(get_visible_value(data_row, 17)),
                            }
                            # S2 is only meaningful on the Mean row; Sensor only on the Min row
                            if colB == "mean":
                                cpt_structured[current_flag]["S2"] = safe_float(get_visible_value(data_row, 21))
                            elif colB == "min":
                                cpt_structured[current_flag]["Sensor"] = safe_float(get_visible_value(data_row, 19))

                    if cpt_structured:
                        parsed_successfully = True
                        st.success("✅ Strategy A successful.")
                except Exception as e:
                    st.write(f"Strategy A failed: {e}")

                # --- STRATEGY B: 2nd Sheet Multi-Row Header Format ---
                if not parsed_successfully:
                    try:
                        df_cpt_alt = pd.read_excel(repo_cpt_file, sheet_name=1, header=None)
                        if df_cpt_alt.iloc[1].astype(str).str.contains("AVG-1").any() == False:
                            row_10 = df_cpt_alt.iloc[10].astype(str).str.strip().str.lower().fillna("")
                            row_11 = df_cpt_alt.iloc[11].astype(str).str.strip().str.lower().fillna("")
                            
                            combined_headers = []
                            for r10, r11 in zip(row_10, row_11):
                                lbl = r11 if r11 and r11 != "nan" else r10
                                lbl = lbl.replace(" ", "").replace("-", "").replace("_", "")
                                if "vc(" in lbl: lbl = "vc"
                                if "sensor(" in lbl: lbl = "sensor"
                                if "%rt" in lbl or "runtime%" in lbl or "runtime" == lbl: lbl = "runtime_pct"
                                combined_headers.append(lbl)
                                
                            df_cpt_alt.columns = combined_headers
                            df_data_rows = df_cpt_alt.iloc[12:].dropna(subset=["datacriteria"]).copy()
                            
                            current_flag = "Unknown"
                            for _, row in df_data_rows.iterrows():
                                val_f1 = str(row.iloc[0]).strip()
                                val_crit = str(row.get("datacriteria", "")).strip().lower()
                                
                                if val_f1 and val_f1 != "nan" and val_f1 != current_flag: current_flag = val_f1
                                if current_flag not in cpt_structured: cpt_structured[current_flag] = {"S2": 0.0, "Sensor": 0.0}
                                    
                                try: rt_val = float(row.get("runtime_pct", 0.0))
                                except (ValueError, TypeError): rt_val = 0.0
                                    
                                if rt_val == 100: continue

                                metric_map = {"mean": "mean", "avg": "mean", "average": "mean",
                                               "min": "min", "max": "max", "(max+min)/2": "(max+min)/2"}
                                metric_key = metric_map.get(val_crit)
                                if metric_key:
                                    cpt_structured[current_flag][metric_key] = {
                                        "tf-1": round(float(row.get("tf1", 0.0)), 1), "tf-2": round(float(row.get("tf2", 0.0)), 1),
                                        "tf-3": round(float(row.get("tf3", 0.0)), 1), "tf-4": round(float(row.get("tf4", 0.0)), 1),
                                        "tf-5": round(float(row.get("tf5", 0.0)), 1), "tc-1": round(float(row.get("tc1", 0.0)), 1),
                                        "tc-2": round(float(row.get("tc2", 0.0)), 1), "tc-3": round(float(row.get("tc3", 0.0)), 1),
                                        "tvc":  round(float(row.get("vc", 0.0)), 1),
                                    }
                                    if metric_key == "mean":
                                        cpt_structured[current_flag]["S2"] = round(float(row.get("s2", 0.0)), 1)
                                    elif metric_key == "min":
                                        cpt_structured[current_flag]["Sensor"] = round(float(row.get("sensor", 0.0)), 1)
                            if cpt_structured: parsed_successfully = True
                    except Exception: pass

                # --- STRATEGY C: Section Layout Matrix Format (Mean values only — this layout has no separate Min/Max/Avg rows) ---
                if not parsed_successfully:
                    try:
                        df_cpt_seg = pd.read_excel(repo_cpt_file, sheet_name=0, header=None)
                        current_flag = "Unknown"
                        for idx, r in df_cpt_seg.iterrows():
                            val_0 = str(r.iloc[0]).strip()
                            if pd.notna(r.iloc[0]) and ("level" in val_0.lower() or "boost" in val_0.lower()):
                                current_flag = val_0
                                if current_flag not in cpt_structured: cpt_structured[current_flag] = {"S2": 0.0, "Sensor": 0.0, "mean": {}}
                                continue
                            if val_0.lower() in ["section", "min", "nan", ""] or pd.isna(r.iloc[0]): continue
                            clean_tag = normalize_sensor_name(val_0)
                            
                            if current_flag != "Unknown":
                                if current_flag not in cpt_structured: cpt_structured[current_flag] = {"S2": 0.0, "Sensor": 0.0, "mean": {}}
                                if clean_tag == "sensor":
                                    try: cpt_structured[current_flag]["Sensor"] = round(float(r.iloc[5]), 1)
                                    except (ValueError, TypeError, IndexError): cpt_structured[current_flag]["Sensor"] = round(float(r.iloc[1]), 1)
                                elif clean_tag == "s2":
                                    cpt_structured[current_flag]["S2"] = round(float(r.iloc[8]), 1)
                                else:
                                    mapping_dict = {
                                        "tf1": "tf-1", "tf2": "tf-2", "tf3": "tf-3", "tf4": "tf-4", "tf5": "tf-5",
                                        "tc1": "tc-1", "tc2": "tc-2", "tc3": "tc-3"
                                    }
                                    if clean_tag in mapping_dict: cpt_structured[current_flag]["mean"][mapping_dict[clean_tag]] = round(float(r.iloc[8]), 1)
                                    elif "tvc" in clean_tag:
                                        if "tvc_vals" not in cpt_structured[current_flag]: cpt_structured[current_flag]["tvc_vals"] = []
                                        cpt_structured[current_flag]["tvc_vals"].append(float(r.iloc[8]))

                        for flg in cpt_structured:
                            if "tvc_vals" in cpt_structured[flg] and cpt_structured[flg]["tvc_vals"]:
                                cpt_structured[flg]["mean"]["tvc"] = round(sum(cpt_structured[flg]["tvc_vals"]) / len(cpt_structured[flg]["tvc_vals"]), 1)
                                del cpt_structured[flg]["tvc_vals"]
                        parsed_successfully = True
                    except Exception: pass

                # 3. SAVE DATA MATRIX AND FORCE RETENTION TO HARD DISK
                if not parsed_successfully or not cpt_structured:
                    raise ValueError("CPT processing pipeline failed. Spreadsheet structural pattern unknown.")

                new_block = {
                    "pulldown_baseline_sensor": resolved_sensor,
                    "original_pulldown_data": copy.deepcopy(p_extracted),
                    "original_cpt_data": copy.deepcopy(cpt_structured),
                    "pulldown_data": copy.deepcopy(p_extracted),
                    "cpt_data": copy.deepcopy(cpt_structured)
                }
                
                verify_db_structure(selected_volume, selected_arrangement, p_repo_key, c_repo_key)
                st.session_state.db[selected_volume][selected_arrangement][p_repo_key][c_repo_key].append(new_block)
                st.session_state.db[selected_volume][selected_arrangement][p_repo_key][c_repo_key] = st.session_state.db[selected_volume][selected_arrangement][p_repo_key][c_repo_key][-10:]
                
                save_memory_to_disk(st.session_state.db)
                
                st.session_state.p_file_key += 1
                st.session_state.cpt_file_key += 1
                
                st.success(f"🚀 Model Simulator Trained successfully! Hard-Backup saved to storage.")
                st.rerun()
            except Exception:
                import traceback
                st.code(traceback.format_exc())
# ================= TAB 3: REVIEWER DASHBOARD =================
with tab3:
    # 1. Secure Authentication Shield Check
    if not st.session_state.reviewer_logged_in:
        st.subheader("🔒 Secure Reviewer Administration Access")
        pass_input = st.text_input("Enter Laboratory Administrative Password:", type="password")
        if st.button("Unlock Admin Dashboard Space", type="primary"):
            if pass_input == REVIEWER_PASSWORD:
                st.session_state.reviewer_logged_in = True
                st.success("Access Granted. Re-routing to matrix editor...")
                st.rerun()
            else:
                st.error("Invalid Administrative Credentials. Access Denied.")
    else:
        # 2. Main Dashboard Layout Once Unlocked
        st.write("### 🔓 Repository Memory Inspection & Manipulation")
        
        c_header1, c_header2 = st.columns([4, 1])
        with c_header1:
            st.write(f"### 🛠️ Data Editor: {selected_volume} | Arrangement: {selected_arrangement}")
        with c_header2:
            if st.button("Close Secure Lock", use_container_width=True):
                st.session_state.reviewer_logged_in = False
                st.rerun()
                
        st.write("---")
        
        # 3. Setup Layout Columns for Inspection Selection
        insp_c1, insp_c2 = st.columns(2)
        
        with insp_c1:
            inspect_p_amb = st.selectbox(
                "Inspect Pulldown Ambient Space:", 
                ["32°C", "43°C"], 
                key="rev_inspect_p_amb"
            )
        with insp_c2:
            inspect_c_amb = st.selectbox(
                "Inspect CPT Ambient Space:", 
                ["16°C", "32°C", "43°C"], 
                index=1,  # Sets default to 32°C so both dropdowns align on load!
                key="rev_inspect_c_amb"
            )
            
        # 4. Dynamic Live Filtering Subtext Status (Fixes your string tracking mismatch error)
        st.caption(f"Currently filtering Pulldown: {inspect_p_amb} / CPT: {inspect_c_amb}")
        
        # 5. Extract Correct Normalized Mapping Dictionary Keys
        p_inspect_key = "32C" if "32" in inspect_p_amb else "43C"
        c_inspect_key = "16C" if "16" in inspect_c_amb else ("32C" if "32" in inspect_c_amb else "43C")
        
        # 6. Safety Verify DB Matrix Sub-Structure Exists
        verify_db_structure(selected_volume, selected_arrangement, p_inspect_key, c_inspect_key)
        records = st.session_state.db[selected_volume][selected_arrangement][p_inspect_key][c_inspect_key]
        
        # 7. Render Active Memory Pool Records Table
        if not records:
            st.info("No paired records matched for this specific arrangement selection.")
        else:
            st.success(f"Found {len(records)} trained datasets stored in hard backup matrix memory loop.")
            
        # --- LOOP THROUGH AND RENDER EACH TRAINED DATASET ---
            for run_idx, record in enumerate(records):
                with st.expander(f"📦 Trained Dataset Record #{run_idx + 1}", expanded=(run_idx == 0)):
                    
                    # Row management buttons
                    c_btn1, c_btn2 = st.columns([4, 1])
                    with c_btn1:
                        st.markdown(f"**Baseline Sensor Target Setting:** `{record.get('pulldown_baseline_sensor', 0.0)}°C`")
                    with c_btn2:
                        if st.button("🗑️ Delete Dataset", key=f"del_ds_{p_inspect_key}_{c_inspect_key}_{run_idx}"):
                            records.pop(run_idx)
                            save_memory_to_disk(st.session_state.db)
                            st.success("Dataset purged successfully!")
                            st.rerun()

                    # Initialize a state version counter for this specific loop record if it doesn't exist
                    version_key = f"ver_{p_inspect_key}_{c_inspect_key}_{run_idx}"
                    if version_key not in st.session_state:
                        st.session_state[version_key] = 1
                    
                    current_ver = st.session_state[version_key]

                    # Section A: Display Pulldown Data Summary Table
                    st.markdown("#### 🔹 Counted Pulldown Matrix")

                    p_df = round_df(add_avg_columns(pd.DataFrame([record["pulldown_data"]])))
                    
                    if "original_pulldown_data" not in record:
                        record["original_pulldown_data"] = record["pulldown_data"].copy()

                    original_p_df = round_df(add_avg_columns(pd.DataFrame([record["original_pulldown_data"]])))

                    def refresh_pulldown():
                        pass

                    # Key dynamically alters based on the current save version string
                    # tf-a / tc-a are computed averages, shown but not directly editable
                    edited_p_df = st.data_editor(
                        p_df,
                        use_container_width=True,
                        hide_index=True,
                        num_rows="fixed",
                        column_config=build_column_config(p_df, disabled_cols=["tf-a", "tc-a"]),
                        key=f"p_edit_{run_idx}_v{current_ver}",
                        on_change=refresh_pulldown
                    )
                    # Recompute averages from whatever tf/tc values were just edited
                    edited_p_df = round_df(add_avg_columns(edited_p_df.drop(columns=["tf-a", "tc-a"], errors="ignore")))
                    
                    st.markdown("##### 📄 Original Uploaded Pulldown Matrix")
                    st.dataframe(
                        original_p_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config=build_column_config(original_p_df)
                    )

                    st.text_area(
                        "📋 Copy Updated Pulldown Matrix",
                        value=edited_p_df.to_csv(sep="\t", index=False),
                        height=180,
                        key=f"p_copy_{run_idx}_v{current_ver}"
                    )
                    
                    # Section B: Display CPT Multivariable Flags Data Matrix
                    st.markdown("#### 🔹 Counted Positions for CPT Matrix")

                    def build_cpt_rows(cpt_data_dict):
                        rows = []
                        for flag_name, flag_block in cpt_data_dict.items():
                            for metric_key in metric_types:
                                metric_data = flag_block.get(metric_key) if isinstance(flag_block, dict) else None
                                if not metric_data:
                                    continue
                                rows.append({
                                    "Test Flag": flag_name,
                                    "Metric": metric_labels[metric_key],
                                    "tf-1": metric_data.get("tf-1", 0.0),
                                    "tf-2": metric_data.get("tf-2", 0.0),
                                    "tf-3": metric_data.get("tf-3", 0.0),
                                    "tf-4": metric_data.get("tf-4", 0.0),
                                    "tf-5": metric_data.get("tf-5", 0.0),
                                    "tc-1": metric_data.get("tc-1", 0.0),
                                    "tc-2": metric_data.get("tc-2", 0.0),
                                    "tc-3": metric_data.get("tc-3", 0.0),
                                    "tvc": metric_data.get("tvc", 0.0),
                                    "S2": flag_block.get("S2", 0.0) if metric_key == "mean" else np.nan,
                                    "Sensor": flag_block.get("Sensor", 0.0) if metric_key == "min" else np.nan,
                                })
                        return rows

                    cpt_rows = build_cpt_rows(record["cpt_data"])

                    if cpt_rows:
                        cpt_df = round_df(add_avg_columns(pd.DataFrame(cpt_rows)))
                        
                        if "original_cpt_data" not in record:
                            record["original_cpt_data"] = record["cpt_data"].copy()

                        original_cpt_rows = build_cpt_rows(record["original_cpt_data"])
                        original_cpt_df = round_df(add_avg_columns(pd.DataFrame(original_cpt_rows)))
                        
                        def refresh_cpt():
                            pass

                        cpt_edit_mode_key = f"cpt_edit_mode_{p_inspect_key}_{c_inspect_key}_{run_idx}"
                        if cpt_edit_mode_key not in st.session_state:
                            st.session_state[cpt_edit_mode_key] = False
                        cpt_editor_widget_key = f"cpt_edit_{run_idx}_v{current_ver}"

                        if not st.session_state[cpt_edit_mode_key]:
                            # VIEW MODE — same real merged-cell (rowspan) look as the Original table below,
                            # reflecting the current (possibly already-edited) values
                            st.markdown(render_merged_cpt_table(cpt_df), unsafe_allow_html=True)
                            if st.button("✏️ Edit CPT Matrix", key=f"cpt_edit_toggle_on_{run_idx}_v{current_ver}"):
                                st.session_state[cpt_edit_mode_key] = True
                                st.rerun()
                            # Nothing pending to save on the CPT side while just viewing
                            edited_cpt_df = cpt_df.copy()
                        else:
                            # EDIT MODE — editable grid (tf-a/tc-a/Test Flag/Metric are locked)
                            edited_cpt_df = st.data_editor(
                                cpt_df,
                                use_container_width=True,
                                hide_index=True,
                                num_rows="fixed",
                                column_config=build_column_config(
                                    cpt_df,
                                    disabled_cols=["tf-a", "tc-a", "Test Flag", "Metric"],
                                    text_cols=["Test Flag", "Metric"]
                                ),
                                key=cpt_editor_widget_key,
                                on_change=refresh_cpt
                            )
                            # Recompute averages from whatever tf/tc values were just edited
                            edited_cpt_df = round_df(add_avg_columns(edited_cpt_df.drop(columns=["tf-a", "tc-a"], errors="ignore")))

                            st.text_area(
                                "📋 Copy Updated CPT Matrix",
                                value=edited_cpt_df.to_csv(sep="\t", index=False),
                                height=220,
                                key=f"cpt_copy_{run_idx}_v{current_ver}"
                            )

                            if st.button("❌ Cancel Editing", key=f"cpt_edit_cancel_{run_idx}_v{current_ver}"):
                                st.session_state.pop(cpt_editor_widget_key, None)
                                st.session_state[cpt_edit_mode_key] = False
                                st.rerun()

                        st.markdown("##### 📄 Original Uploaded CPT Matrix")
                        # Always reflects the file as first uploaded — never touched by edits
                        st.markdown(render_merged_cpt_table(original_cpt_df), unsafe_allow_html=True)

                        # Detect changes
                        pulldown_changed = not edited_p_df.equals(p_df)
                        cpt_changed = not edited_cpt_df.equals(cpt_df)
                        dataset_changed = pulldown_changed or cpt_changed



                        if dataset_changed:
                            st.success("🟡 Unsaved changes detected.")
                            save_clicked = st.button(
                                "💾 Save Edited Dataset",
                                key=f"save_dataset_{run_idx}_v{current_ver}",
                                type="primary"
                            )
                        else:
                            st.info("No changes made.")
                            save_clicked = False

                        if save_clicked:
                            # Save Pulldown Matrix (tf-a/tc-a are computed display-only fields, never stored)
                            record["pulldown_data"] = edited_p_df.drop(columns=["tf-a", "tc-a"], errors="ignore").iloc[0].to_dict()

                            # Save CPT Matrix — reassemble the 4 rows per flag back into the nested structure
                            label_to_key = {v: k for k, v in metric_labels.items()}
                            new_cpt = {}
                            for _, row in edited_cpt_df.iterrows():
                                flag = row["Test Flag"]
                                metric_key = label_to_key.get(row["Metric"], str(row["Metric"]).lower())

                                if flag not in new_cpt:
                                    new_cpt[flag] = {"S2": 0.0, "Sensor": 0.0}

                                new_cpt[flag][metric_key] = {
                                    "tf-1": float(row["tf-1"]),
                                    "tf-2": float(row["tf-2"]),
                                    "tf-3": float(row["tf-3"]),
                                    "tf-4": float(row["tf-4"]),
                                    "tf-5": float(row["tf-5"]),
                                    "tc-1": float(row["tc-1"]),
                                    "tc-2": float(row["tc-2"]),
                                    "tc-3": float(row["tc-3"]),
                                    "tvc": float(row["tvc"]),
                                }
                                if metric_key == "mean" and pd.notna(row["S2"]):
                                    new_cpt[flag]["S2"] = float(row["S2"])
                                if metric_key == "min" and pd.notna(row["Sensor"]):
                                    new_cpt[flag]["Sensor"] = float(row["Sensor"])
                            record["cpt_data"] = new_cpt

                            # Commit file data structure changes to the physical disk 
                            save_memory_to_disk(st.session_state.db)
                            
                            # Return the CPT matrix to its merged read-only view, now showing the saved values
                            st.session_state[cpt_edit_mode_key] = False

                            # INCREMENT VERSION: Wipes out the stale data cache instantly on rerun
                            st.session_state[version_key] += 1
                            
                            st.success("✅ Dataset updated successfully.")
                            st.rerun()
                    else:
                        st.warning("No CPT entries found inside this specific record block.")
